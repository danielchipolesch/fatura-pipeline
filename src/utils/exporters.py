"""
Exportadores tabulares do pipeline de faturas.

Gera e mantém dois arquivos acumulativos no diretório de saída:
  - output/xlsx/faturas.xlsx   (uma linha por item de fatura)
  - output/csv/faturas.csv     (mesma estrutura, UTF-8 BOM para compatibilidade Excel)

Estrutura das linhas:
  Campos da fatura (identificação, datas, partes, totais, tributos)
  repetidos para cada item de linha — facilita filtros e tabelas dinâmicas.
  Quando a fatura não possui itens, uma única linha é gerada com os campos
  de item em branco.
"""
from __future__ import annotations

import csv
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from loguru import logger

# ---------------------------------------------------------------------------
# Colunas (ordem canônica)
# ---------------------------------------------------------------------------

COLUMNS: list[str] = [
    # Identificação da fatura
    "source_file",
    "invoice_number",
    "invoice_layout",
    "tipo_fatura_operacional",
    "access_key",
    # Datas
    "issue_date",
    "due_date",
    "service_period_start",
    "service_period_end",
    # Fornecedor
    "supplier_name",
    "supplier_cnpj",
    "supplier_ie",
    "supplier_state",
    # Cliente
    "customer_name",
    "customer_cnpj",
    "customer_cpf",
    "customer_state",
    # Instalação de energia
    "consumer_unit",
    "client_number",
    # Item de linha
    "item_code",
    "item_description",
    "item_unit",
    "item_quantity",
    "item_unit_price",
    "item_discount",
    "item_total",
    # Totais da fatura
    "subtotal",
    "discount",
    "total_taxes",
    "total",
    "currency",
    # Tributos (ICMS / PIS / COFINS)
    "icms_base",
    "icms_rate",
    "icms_amount",
    "pis_base",
    "pis_rate",
    "pis_amount",
    "cofins_base",
    "cofins_rate",
    "cofins_amount",
    "outros_tributos",
    # Pagamento e metadados
    "bank_slip_line",
    "parsing_method",
    "confidence_score",
]

# Cabeçalhos legíveis para exibição no Excel
_HEADERS: dict[str, str] = {
    "source_file": "Arquivo PDF",
    "invoice_number": "Número da NF",
    "invoice_layout": "Layout",
    "tipo_fatura_operacional": "Tipo Operacional",
    "access_key": "Chave de Acesso",
    "issue_date": "Data de Emissão",
    "due_date": "Vencimento",
    "service_period_start": "Competência Início",
    "service_period_end": "Competência Fim",
    "supplier_name": "Fornecedor",
    "supplier_cnpj": "CNPJ Fornecedor",
    "supplier_ie": "IE Fornecedor",
    "supplier_state": "UF Fornecedor",
    "customer_name": "Cliente",
    "customer_cnpj": "CNPJ Cliente",
    "customer_cpf": "CPF Cliente",
    "customer_state": "UF Cliente",
    "consumer_unit": "Unidade Consumidora",
    "client_number": "Número do Cliente",
    "item_code": "Cód. Item",
    "item_description": "Descrição do Item",
    "item_unit": "Unidade",
    "item_quantity": "Quantidade",
    "item_unit_price": "Preço Unitário",
    "item_discount": "Desconto Item",
    "item_total": "Valor Item",
    "subtotal": "Subtotal",
    "discount": "Desconto Fatura",
    "total_taxes": "Total Tributos",
    "total": "Total da Fatura",
    "currency": "Moeda",
    "icms_base": "Base ICMS",
    "icms_rate": "Alíquota ICMS (%)",
    "icms_amount": "ICMS",
    "pis_base": "Base PIS",
    "pis_rate": "Alíquota PIS (%)",
    "pis_amount": "PIS",
    "cofins_base": "Base COFINS",
    "cofins_rate": "Alíquota COFINS (%)",
    "cofins_amount": "COFINS",
    "outros_tributos": "Outros Tributos (JSON)",
    "bank_slip_line": "Linha Digitável",
    "parsing_method": "Método de Extração",
    "confidence_score": "Score de Confiança",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

# Colunas principais para ajuste automático de largura (evita calcular todas)
_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 60


# ---------------------------------------------------------------------------
# Achatamento do dicionário de fatura
# ---------------------------------------------------------------------------

def _safe(value: Any) -> Any:
    """Normaliza valores para células de planilha: mantém None, str, int, float, date."""
    if value is None:
        return None
    if isinstance(value, (bool, int, float, date, datetime)):
        return value
    if isinstance(value, str):
        # READ_ERROR_SENTINEL e outros sentinelas ficam como string
        return value
    # Listas/dicts não esperados → JSON compacto
    return json.dumps(value, ensure_ascii=False)


def invoice_to_rows(invoice_dict: dict) -> list[dict]:
    """
    Converte um dicionário de fatura em uma lista de linhas planas.
    Retorna uma linha por item de fatura; se não houver itens, retorna
    uma única linha com os campos de item nulos.
    """
    supplier = invoice_dict.get("supplier") or {}
    supplier_addr = supplier.get("address") or {}
    customer = invoice_dict.get("customer") or {}
    customer_addr = customer.get("address") or {}

    # Campos de tributos: ICMS / PIS / COFINS em colunas dedicadas
    taxes: list[dict] = invoice_dict.get("taxes") or []
    tax_map: dict[str, dict] = {}
    outros: list[dict] = []
    for t in taxes:
        name = (t.get("name") or "").upper()
        if name in ("ICMS", "PIS", "COFINS"):
            tax_map[name] = t
        else:
            outros.append(t)

    def _tax(name: str, field: str) -> Any:
        return _safe(tax_map.get(name, {}).get(field))

    base: dict[str, Any] = {
        "source_file": _safe(invoice_dict.get("source_file")),
        "invoice_number": _safe(invoice_dict.get("invoice_number")),
        "invoice_layout": _safe(invoice_dict.get("invoice_layout")),
        "tipo_fatura_operacional": _safe(invoice_dict.get("tipo_fatura_operacional")),
        "access_key": _safe(invoice_dict.get("access_key")),
        "issue_date": _safe(invoice_dict.get("issue_date")),
        "due_date": _safe(invoice_dict.get("due_date")),
        "service_period_start": _safe(invoice_dict.get("service_period_start")),
        "service_period_end": _safe(invoice_dict.get("service_period_end")),
        "supplier_name": _safe(supplier.get("name")),
        "supplier_cnpj": _safe(supplier.get("cnpj")),
        "supplier_ie": _safe(supplier.get("ie")),
        "supplier_state": _safe(supplier_addr.get("state")),
        "customer_name": _safe(customer.get("name")),
        "customer_cnpj": _safe(customer.get("cnpj")),
        "customer_cpf": _safe(customer.get("cpf")),
        "customer_state": _safe(customer_addr.get("state")),
        "consumer_unit": _safe(invoice_dict.get("consumer_unit")),
        "client_number": _safe(invoice_dict.get("client_number")),
        "subtotal": _safe(invoice_dict.get("subtotal")),
        "discount": _safe(invoice_dict.get("discount")),
        "total_taxes": _safe(invoice_dict.get("total_taxes")),
        "total": _safe(invoice_dict.get("total")),
        "currency": _safe(invoice_dict.get("currency")),
        "icms_base": _tax("ICMS", "base"),
        "icms_rate": _tax("ICMS", "rate"),
        "icms_amount": _tax("ICMS", "amount"),
        "pis_base": _tax("PIS", "base"),
        "pis_rate": _tax("PIS", "rate"),
        "pis_amount": _tax("PIS", "amount"),
        "cofins_base": _tax("COFINS", "base"),
        "cofins_rate": _tax("COFINS", "rate"),
        "cofins_amount": _tax("COFINS", "amount"),
        "outros_tributos": json.dumps(outros, ensure_ascii=False) if outros else None,
        "bank_slip_line": _safe(invoice_dict.get("bank_slip_line")),
        "parsing_method": _safe(invoice_dict.get("parsing_method")),
        "confidence_score": _safe(invoice_dict.get("confidence_score")),
    }

    line_items: list[dict] = invoice_dict.get("line_items") or []
    if not line_items:
        row = dict(base)
        for col in ("item_code", "item_description", "item_unit",
                    "item_quantity", "item_unit_price", "item_discount", "item_total"):
            row[col] = None
        return [row]

    rows: list[dict] = []
    for item in line_items:
        row = dict(base)
        row["item_code"] = _safe(item.get("code"))
        row["item_description"] = _safe(item.get("description"))
        row["item_unit"] = _safe(item.get("unit"))
        row["item_quantity"] = _safe(item.get("quantity"))
        row["item_unit_price"] = _safe(item.get("unit_price"))
        row["item_discount"] = _safe(item.get("discount"))
        row["item_total"] = _safe(item.get("total"))
        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Escrita / append XLSX
# ---------------------------------------------------------------------------

def _write_header_row(ws: Any) -> None:
    """Escreve a linha de cabeçalho com estilo e congela o painel."""
    header_values = [_HEADERS.get(c, c) for c in COLUMNS]
    ws.append(header_values)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"


def _auto_width(ws: Any) -> None:
    """Ajusta largura das colunas com base no conteúdo atual."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH
        )


def append_to_xlsx(rows: list[dict], xlsx_path: Path) -> None:
    """Acrescenta linhas ao arquivo XLSX acumulativo. Cria o arquivo se não existir."""
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_path.exists():
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
        except Exception as exc:
            logger.warning(f"Não foi possível abrir {xlsx_path.name}, recriando: {exc}")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Faturas"
            _write_header_row(ws)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Faturas"
        _write_header_row(ws)

    for row in rows:
        ws.append([row.get(c) for c in COLUMNS])

    _auto_width(ws)
    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# Escrita / append CSV
# ---------------------------------------------------------------------------

def append_to_csv(rows: list[dict], csv_path: Path) -> None:
    """Acrescenta linhas ao arquivo CSV acumulativo. Cria com cabeçalho se não existir."""
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not csv_path.exists()

    with open(csv_path, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=COLUMNS,
            extrasaction="ignore",
            quoting=csv.QUOTE_MINIMAL,
        )
        if write_header:
            # Usa cabeçalhos legíveis também no CSV
            writer.writerow({c: _HEADERS.get(c, c) for c in COLUMNS})
        for row in rows:
            # Converte tudo para string simples no CSV
            writer.writerow({
                c: ("" if row.get(c) is None else str(row[c]))
                for c in COLUMNS
            })


# ---------------------------------------------------------------------------
# Ponto de entrada único
# ---------------------------------------------------------------------------

def export_invoice(invoice_dict: dict, xlsx_path: Path, csv_path: Path) -> None:
    """
    Exporta uma fatura para os arquivos acumulativos XLSX e CSV.
    Falhas de exportação são logadas como warning sem interromper o pipeline.
    """
    try:
        rows = invoice_to_rows(invoice_dict)
        append_to_xlsx(rows, xlsx_path)
        append_to_csv(rows, csv_path)
        n = len(rows)
        logger.debug(
            f"Exportado: {invoice_dict.get('source_file')} → "
            f"{n} linha(s) adicionada(s) a XLSX e CSV"
        )
    except Exception as exc:
        logger.warning(
            f"Falha ao exportar '{invoice_dict.get('source_file')}' "
            f"para XLSX/CSV: {exc}"
        )
